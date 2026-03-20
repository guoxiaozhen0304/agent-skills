#!/usr/bin/env python3
"""
Generate Excel test case file from YAML test case file.

Usage:
    python generate_excel.py <input.yaml> [output.xlsx] [--template /path/to/Template.xlsx]

Template search order (first match wins):
    1. --template argument (explicit path)
    2. Same directory as the yaml file: Template.xlsx
    3. Parent directories of the yaml file (up to 3 levels): Template.xlsx
    4. Current working directory: Template.xlsx
    5. Fallback: built-in style (original behavior)
"""

import sys
import os
import shutil
import copy
import yaml
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system("%s -m pip install openpyxl pyyaml -q" % sys.executable)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ---------- Fallback style constants (used when no template found) ----------
COLOR_HEADER_BG = "2F5496"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ALT_ROW   = "EEF4FF"
COLOR_BORDER    = "B8C4D8"

PRIORITY_COLORS = {
    "高": "C00000",
    "中": "E07000",
    "低": "375623",
    "P0": "C00000",
    "P1": "C00000",
    "P2": "E07000",
    "P3": "375623",
}

FALLBACK_COLUMNS = [
    ("用例ID",        10),
    ("模块",          14),
    ("用例名称",      30),
    ("用例类型",      14),
    ("优先级",        10),
    ("是否可自动化",  12),
    ("前置条件",      30),
    ("测试步骤",      45),
    ("预期结果",      35),
    ("实际结果",      25),
    ("执行状态",      12),
    ("备注",          20),
]

AUTOMATABLE_COLORS = {
    "是":   "375623",   # 深绿
    "否":   "C00000",   # 深红
    "部分": "E07000",   # 橙色
}

# Template column mapping: column_letter -> (yaml_field_or_callable, default_value)
# Matches Template.xlsx structure:
#   A:*用例名称  B:用例编号  C:处理者  D:用例状态  E:用例类型  F:用例等级
#   G:迭代  H:模块  I:需求编号  J:缺陷编号  K:描述  L:前置条件
#   M:归属目录  N:标签  O:自定义字段1  P:自定义字段2  Q:测试步骤模式
#   R:测试步骤1  S:预期结果1  T:测试步骤2  U:预期结果2 ... AD:测试步骤7  AE:预期结果7
TEMPLATE_COLUMN_MAP = {
    "A": "name",
    "B": "id",
    "C": "",           # 处理者 - empty
    "D": "status",
    "E": "type",
    "F": "priority",   # will be mapped to Priority 1/2/3
    "G": "",           # 迭代 - empty
    "H": "module",
    "I": "",           # 需求编号 - empty
    "J": "",           # 缺陷编号 - empty
    "K": "expected_result",
    "L": "preconditions",
    "M": "module",     # 归属目录 = module
    "N": "type",       # 标签 = type
    "O": "automatable",  # 自定义字段1 = 是否可自动化
    "P": "",
    "Q": "__steps_mode__",  # always "步骤"
    # R-AE: steps 1-7 (action/expected pairs)
}

PRIORITY_MAP = {
    "高":  "Priority 1",
    "中":  "Priority 2",
    "低":  "Priority 3",
    "P0":  "Priority 0",
    "P1":  "Priority 1",
    "P2":  "Priority 2",
    "P3":  "Priority 3",
    "P4":  "Priority 4",
}

MAX_STEPS = 7  # Template supports 7 steps (columns R-AE)


# ---------- Helpers ----------

def thin_border():
    side = Side(style="thin", color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def format_list(value) -> str:
    if not value:
        return ""
    if isinstance(value, list):
        return "\n".join("- %s" % item for item in value)
    return str(value)


def format_steps_fallback(steps) -> str:
    """Format steps as single string for fallback (non-template) mode."""
    if not steps:
        return ""
    if isinstance(steps, str):
        return steps
    lines = []
    for i, step in enumerate(steps, start=1):
        if isinstance(step, dict):
            action = step.get("action", step.get("step", ""))
            expected = step.get("expected", "")
            line = "%d. %s" % (i, action)
            if expected:
                line += "\n   -> 期望：%s" % expected
        else:
            line = "%d. %s" % (i, step)
        lines.append(line)
    return "\n".join(lines)


def parse_steps(steps):
    """Parse steps into list of (action, expected) tuples, max MAX_STEPS."""
    if not steps:
        return []
    if isinstance(steps, str):
        return [(steps, "")]
    result = []
    for step in steps[:MAX_STEPS]:
        if isinstance(step, dict):
            action = step.get("action", step.get("step", ""))
            expected = step.get("expected", "")
        else:
            action = str(step)
            expected = ""
        result.append((action, expected))
    return result


def load_yaml(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def make_header_style(size=11):
    return (
        Font(name="Microsoft YaHei", bold=True, color=COLOR_HEADER_FG, size=size),
        PatternFill("solid", fgColor=COLOR_HEADER_BG),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
    )


def find_template(yaml_path: str, explicit_template: str = None) -> str:
    """Search for Template.xlsx, return path if found, else None."""
    if explicit_template and os.path.isfile(explicit_template):
        return explicit_template

    candidates = []
    # 1. Same dir as yaml
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(yaml_path)), "Template.xlsx"))
    # 2. Parent directories (up to 3 levels)
    d = os.path.dirname(os.path.abspath(yaml_path))
    for _ in range(3):
        d = os.path.dirname(d)
        candidates.append(os.path.join(d, "Template.xlsx"))
    # 3. Current working directory
    candidates.append(os.path.join(os.getcwd(), "Template.xlsx"))

    for path in candidates:
        if os.path.isfile(path):
            return path
    return None


def copy_cell_style(src_cell, dst_cell):
    """Copy style attributes from src_cell to dst_cell."""
    if src_cell.font:
        dst_cell.font = copy.copy(src_cell.font)
    if src_cell.fill:
        dst_cell.fill = copy.copy(src_cell.fill)
    if src_cell.border:
        dst_cell.border = copy.copy(src_cell.border)
    if src_cell.alignment:
        dst_cell.alignment = copy.copy(src_cell.alignment)


# ---------- Template-based generation ----------

def generate_with_template(wb_template, test_cases: list, suite_name: str, output_path: str):
    """Write test cases into a copy of the template workbook."""
    # Work on the first sheet (sheet1)
    ws = wb_template.worksheets[0]

    # Delete all rows except header (row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Capture sample row style from original template row 2 for reuse
    # (already deleted - we'll use a default data cell style instead)
    data_font = Font(name="宋体", size=12)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row_idx, tc_data in enumerate(test_cases):
        row_num = row_idx + 2  # data starts at row 2

        priority_raw = str(tc_data.get("priority", "中"))
        priority_val = PRIORITY_MAP.get(priority_raw, priority_raw)
        precond_str  = format_list(tc_data.get("preconditions", tc_data.get("precondition", "")))
        steps_parsed = parse_steps(tc_data.get("steps", tc_data.get("test_steps", [])))

        # Build column-to-value mapping
        col_values = {}
        for col_letter, field in TEMPLATE_COLUMN_MAP.items():
            if not field:
                col_values[col_letter] = ""
            elif field == "priority":
                col_values[col_letter] = priority_val
            elif field == "preconditions":
                col_values[col_letter] = precond_str
            elif field == "expected_result":
                col_values[col_letter] = tc_data.get("expected_result", tc_data.get("expected", ""))
            elif field == "status":
                col_values[col_letter] = tc_data.get("status", "未执行")
            elif field == "__steps_mode__":
                col_values[col_letter] = "步骤" if steps_parsed else ""
            else:
                col_values[col_letter] = tc_data.get(field, tc_data.get("title" if field == "name" else field, ""))

        # Steps: columns R/S, T/U, V/W, X/Y, Z/AA, AB/AC, AD/AE
        step_col_pairs = [
            ("R", "S"), ("T", "U"), ("V", "W"),
            ("X", "Y"), ("Z", "AA"), ("AB", "AC"), ("AD", "AE"),
        ]
        for i, (act_col, exp_col) in enumerate(step_col_pairs):
            if i < len(steps_parsed):
                col_values[act_col] = steps_parsed[i][0]
                col_values[exp_col] = steps_parsed[i][1]
            else:
                col_values[act_col] = ""
                col_values[exp_col] = ""

        # Write cells
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            value = col_values.get(col_letter, "")
            cell = ws.cell(row=row_num, column=col_idx, value=str(value) if value else "")
            cell.font = data_font
            # Left-align text-heavy columns
            if col_letter in ("A", "K", "L", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"):
                cell.alignment = data_align_left
            else:
                cell.alignment = data_align

        # Adjust row height based on content
        step_count = len(steps_parsed)
        ws.row_dimensions[row_num].height = max(20, step_count * 18)

    # Add summary sheet
    if "统计汇总" in wb_template.sheetnames:
        del wb_template["统计汇总"]
    ws_summary = wb_template.create_sheet("统计汇总")
    _write_summary(ws_summary, test_cases, suite_name)

    wb_template.save(output_path)
    print("[OK] Excel generated (template mode): %s" % output_path)
    print("     Total cases: %d" % len(test_cases))
    print("     Template columns: %d" % ws.max_column)
    return output_path


# ---------- Fallback (built-in style) generation ----------

def generate_fallback(test_cases: list, suite_name: str, output_path: str):
    """Original built-in style generation."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    h_font, h_fill, h_align = make_header_style(13)

    title_text = "%s   生成时间：%s" % (suite_name, datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FALLBACK_COLUMNS))
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = h_font
    tc.fill = h_fill
    tc.alignment = h_align
    ws.row_dimensions[1].height = 32

    col_font, col_fill, col_align = make_header_style(11)
    for col_idx, (col_name, col_width) in enumerate(FALLBACK_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = col_font
        cell.fill = col_fill
        cell.alignment = col_align
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    wrap_top = Alignment(vertical="top", wrap_text=True)
    alt_fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    for i, tc_data in enumerate(test_cases):
        row_num = i + 3
        is_alt = (i % 2 == 1)

        steps_str    = format_steps_fallback(tc_data.get("steps", tc_data.get("test_steps", "")))
        precond_str  = format_list(tc_data.get("preconditions", tc_data.get("precondition", "")))
        priority     = str(tc_data.get("priority", "中"))
        automatable  = str(tc_data.get("automatable", ""))

        row_values = [
            tc_data.get("id", ""),
            tc_data.get("module", tc_data.get("feature", "")),
            tc_data.get("name", tc_data.get("title", "")),
            tc_data.get("type", tc_data.get("test_type", "功能测试")),
            priority,
            automatable,
            precond_str,
            steps_str,
            tc_data.get("expected_result", tc_data.get("expected", "")),
            tc_data.get("actual_result", ""),
            tc_data.get("status", "未执行"),
            tc_data.get("remarks", tc_data.get("notes", "")),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=str(value) if value else "")
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = wrap_top
            cell.border = thin_border()
            if is_alt:
                cell.fill = alt_fill

        if priority in PRIORITY_COLORS:
            pri_cell = ws.cell(row=row_num, column=5)
            pri_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=PRIORITY_COLORS[priority])

        if automatable in AUTOMATABLE_COLORS:
            auto_cell = ws.cell(row=row_num, column=6)
            auto_cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=AUTOMATABLE_COLORS[automatable])

        step_lines = steps_str.count("\n") + 1 if steps_str else 1
        ws.row_dimensions[row_num].height = max(20, step_lines * 15)

    ws2 = wb.create_sheet("统计汇总")
    _write_summary(ws2, test_cases, suite_name)

    wb.save(output_path)
    print("[OK] Excel generated (built-in style): %s" % output_path)
    print("     Total cases: %d" % len(test_cases))
    return output_path


# ---------- Summary sheet ----------

def _write_summary(ws, test_cases: list, suite_name: str):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12

    h_font, h_fill, _ = make_header_style(11)
    center = Alignment(horizontal="center", vertical="center")
    border = thin_border()

    def hdr(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = h_font
        c.fill = h_fill
        c.alignment = center
        c.border = border
        ws.row_dimensions[row].height = 24

    def val_cell(row, col, v):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(name="Microsoft YaHei", size=11)
        c.alignment = center
        c.border = border

    title_cell = ws.cell(row=1, column=1, value="测试套件：%s" % suite_name)
    title_cell.font = Font(name="Microsoft YaHei", bold=True, size=13)
    ws.row_dimensions[1].height = 28

    priorities, types, modules = {}, {}, {}
    automatable_stats = {}
    for tc in test_cases:
        p = tc.get("priority", "未知")
        priorities[p] = priorities.get(p, 0) + 1
        t = tc.get("type", tc.get("test_type", "未知"))
        types[t] = types.get(t, 0) + 1
        m = tc.get("module", tc.get("feature", "未知"))
        modules[m] = modules.get(m, 0) + 1
        a = tc.get("automatable", "未标注")
        automatable_stats[a] = automatable_stats.get(a, 0) + 1

    r = 3
    hdr(r, 1, "优先级"); hdr(r, 2, "数量"); r += 1
    for p, cnt in sorted(priorities.items()):
        val_cell(r, 1, p); val_cell(r, 2, cnt); r += 1
    val_cell(r, 1, "合计"); val_cell(r, 2, len(test_cases)); r += 2

    hdr(r, 1, "用例类型"); hdr(r, 2, "数量"); r += 1
    for t, cnt in sorted(types.items()):
        val_cell(r, 1, t); val_cell(r, 2, cnt); r += 1
    r += 1

    hdr(r, 1, "是否可自动化"); hdr(r, 2, "数量"); r += 1
    for a, cnt in sorted(automatable_stats.items()):
        val_cell(r, 1, a); val_cell(r, 2, cnt); r += 1
    r += 1

    hdr(r, 1, "模块"); hdr(r, 2, "数量"); r += 1
    for m, cnt in sorted(modules.items()):
        val_cell(r, 1, m); val_cell(r, 2, cnt); r += 1


# ---------- Entry point ----------

def generate_excel(yaml_path: str, output_path: str = None, template_path: str = None):
    data = load_yaml(yaml_path)

    if isinstance(data, list):
        test_cases = data
        suite_name = "测试用例集"
    else:
        test_cases = data.get("test_cases", data.get("cases", []))
        suite_name = (
            data.get("test_suite", {}).get("name", "")
            or data.get("name", "测试用例集")
        )

    if not output_path:
        base = os.path.splitext(yaml_path)[0]
        output_path = base + ".xlsx"

    # Search for template
    found_template = find_template(yaml_path, template_path)

    if found_template:
        print("[INFO] Using template: %s" % found_template)
        # Copy template to a temp location to avoid modifying the original
        tmp_path = output_path + ".tmp.xlsx"
        shutil.copy2(found_template, tmp_path)
        try:
            wb = openpyxl.load_workbook(tmp_path)
            result = generate_with_template(wb, test_cases, suite_name, output_path)
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        return result
    else:
        print("[INFO] Template.xlsx not found, using built-in style")
        return generate_fallback(test_cases, suite_name, output_path)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <input.yaml> [output.xlsx] [--template path]")
        sys.exit(1)

    yaml_path = sys.argv[1]
    out_path = None
    tpl_path = None

    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == "--template" and i + 1 < len(sys.argv):
            tpl_path = sys.argv[i + 1]
            i += 2
        elif not out_path and not sys.argv[i].startswith("--"):
            out_path = sys.argv[i]
            i += 1
        else:
            i += 1

    generate_excel(yaml_path, out_path, tpl_path)
